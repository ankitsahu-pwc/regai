"""BRD/FRD to regulatory readiness questionnaire generator.

Historical role
---------------
This module used to be a fully-deterministic template pipeline: it derived
impact pairs from a BRD, then synthesised a bank of questions from
hardcoded option families, theme dictionaries and generic prompts. The
questions were scored via a fixed answer-to-score lookup table.

Current role (AI-agent enhancement)
-----------------------------------
All hardcoded questionnaire values, static option sets, fixed scoring values
and generic question templates have been removed. Question generation is now
produced by :mod:`services.ai_questionnaire_generator`, which prompts the
GenAI Shared Service with the full regulatory context (obligations, BRD
requirements, RTM entries, impact assessment, selected client roles) and
returns adaptive, per-question option sets with option-specific follow-ups
and per-option scoring.

What remains in this module:

* :class:`Requirement`, :class:`ImpactPair`, :class:`Question` dataclasses
  (contract for the rest of the pipeline).
* BRD parsing helpers (:func:`read_docx_requirements`,
  :func:`requirements_from_report`).
* Impact-pair derivation helpers (:func:`derive_impact_pairs` and the
  supporting ``AREA_KEYWORDS`` / ``FUNCTION_KEYWORDS`` taxonomies — these
  are analytical utilities used by Agent 1 and 3 to bucket requirements,
  NOT question templates).
* Validation, scoring metadata, package assembly, dedup and the Excel
  writer.
* Backwards-compatible entry points (:func:`build_package_from_report`,
  :func:`build_questionnaire_package`) that now call the AI generator
  with any additional context the caller supplies.

When the caller passes a live :class:`~services.genai_service.GenAIClient`
via the new context-aware entry points the questionnaire is fully AI
generated. When no client is available the AI generator emits a small set
of "Manual Review Required" placeholders per impact pair — it never
fabricates hardcoded questions.
"""

from __future__ import annotations

import logging
import os
import re as _re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple, Union

from openpyxl import Workbook

logger = logging.getLogger(__name__)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from utils.docx_parser import (
    DocxSource,
    clean_text,
    iter_sectioned_tables,
    normalise_header,
)

# ---------------------------------------------------------------------------
# Tunables (env-overridable, identical defaults to v11)
# ---------------------------------------------------------------------------

CONFIDENCE_FLOOR = int(os.getenv("QUESTION_CONFIDENCE_FLOOR", "90"))
# v13.1 — the package-confidence metric is now a **content-correctness** grade,
# so the historic 90 floor has been retired by default. Set
# OVERALL_QUESTIONNAIRE_CONFIDENCE_FLOOR=90 in the environment to restore the
# legacy behaviour. Set PACKAGE_CONFIDENCE_MODE=structural to revert the
# whole metric to the v11/v12 structural-completeness grade.
OVERALL_CONFIDENCE_FLOOR = int(os.getenv("OVERALL_QUESTIONNAIRE_CONFIDENCE_FLOOR", "0"))
PACKAGE_CONFIDENCE_MODE = os.getenv("PACKAGE_CONFIDENCE_MODE", "content").strip().lower()
MAX_AREA_FUNCTION_PAIRS = int(os.getenv("MAX_AREA_FUNCTION_PAIRS", "40"))
MIN_FREE_TEXT = int(os.getenv("MIN_FREE_TEXT_QUESTIONS", "5"))
MAX_FREE_TEXT = int(os.getenv("MAX_FREE_TEXT_QUESTIONS", "10"))

# v13.1 — the keys that must be populated in every question's
# ``explainability`` bundle for the content-correctness grade to credit it
# as a fully-traceable question.
EXPLAINABILITY_REQUIRED_KEYS = (
    "regulation",
    "regulator",
    "article",
    "obligation_id",
    "brd_requirement_ids",
    "business_function",
    "control_objective",
    "reason",
    "expected_evidence",
    "risk_if_negative",
)

# ---------------------------------------------------------------------------
# Regulatory taxonomy (lifted verbatim)
# ---------------------------------------------------------------------------

REGULATORY_TAXONOMY = {
    "DORA": {
        "official_source": "Regulation (EU) 2022/2554 and related DORA RTS/ITS guidance",
        "pillars": [
            "Governance and organisation",
            "ICT risk management framework",
            "ICT systems, protocols and tools",
            "Identification, protection, prevention, detection, response and recovery",
            "Backup, restoration, recovery and communication",
            "Incident management, classification and reporting",
            "Digital operational resilience testing",
            "ICT third-party risk management",
            "Key contractual provisions and exit planning",
            "Information security, access control, data protection and auditability",
            "Management reporting and evidence traceability",
        ],
        "article_hints": {
            "governance": "DORA Article 5",
            "framework": "DORA Article 6",
            "identification": "DORA Article 8",
            "protection": "DORA Article 9",
            "detection": "DORA Article 10",
            "response": "DORA Article 11",
            "backup": "DORA Article 12",
            "incident": "DORA Articles 17-20",
            "testing": "DORA Articles 24-27",
            "third_party": "DORA Articles 28-30",
            "contracts": "DORA Article 30",
        },
    }
}

AREA_KEYWORDS = {
    "Front Office": ["front office", "client", "customer", "trading", "sales", "market"],
    "Middle Office": ["middle office", "risk control", "valuation", "trade support", "risk"],
    "Back Office": ["back office", "settlement", "reconciliation", "processing", "operations"],
    "Regulatory Reporting & Financial Reporting": ["report", "dashboard", "governance pack", "management body", "kpi", "kri"],
    "Business Structure & Functions": ["business function", "critical function", "important function", "service mapping", "institution"],
    "Firm Type / Client Type": ["tier", "entity", "financial services", "proportionality", "institution"],
    "Operating Model": ["operating model", "workflow", "role", "responsibil", "owner", "attestation"],
    "Risk & Controls framework": ["risk", "control", "framework", "vulnerability", "security", "risk acceptance"],
    "Governance Model": ["governance", "management body", "board", "approval", "decision"],
    "Internal Compliances": ["compliance", "audit", "policy", "procedure", "legal"],
    "Third Party Risk Management / Dependency": ["third-party", "third party", "vendor", "provider", "subcontract", "contract", "exit"],
    "Programme Maturity / Programme Ownership": ["programme", "maturity", "readiness", "roadmap", "phase", "implementation"],
    "Program Sponsorship / Budget Planning": ["budget", "sponsor", "resource", "funding", "buying", "support"],
    "People, Policies & Processes": ["training", "people", "policy", "process", "procedure", "lessons learned"],
    "IT, Systems & Technology": ["ict", "system", "application", "infrastructure", "cmdb", "itsm", "siem", "iam"],
    "Data Reporting & Governance": ["data", "metadata", "lineage", "quality", "evidence", "timestamp", "dictionary"],
    "IT Security / Cyber Security": ["security", "cyber", "access", "privileged", "vulnerability", "encryption", "intrusion"],
    "Legal (Contracts, Readiness & Agreements)": ["legal", "contract", "clause", "termination", "audit rights", "access rights", "jurisdiction"],
    "HR": ["hr", "training", "resource", "sme", "people"],
    "High Impact Pain Points": ["gap", "pain", "manual", "fragment", "incomplete", "overdue", "unclear", "inconsistent", "constraint"],
}

FUNCTION_KEYWORDS = {
    "Execution / Client Activity": ["front office", "client", "trading", "customer", "market"],
    "Risk Management": ["risk", "control", "risk appetite", "risk assessment", "vulnerability"],
    "Compliance & Legal": ["compliance", "legal", "policy", "regulatory", "contract", "audit"],
    "Technology / IT Operations": ["ict", "system", "application", "infrastructure", "cmdb", "itsm", "backup", "restore"],
    "Cyber Security": ["security", "cyber", "siem", "vulnerability", "access", "encryption"],
    "Business Continuity / Resilience": ["continuity", "recovery", "resilience", "testing", "backup", "restore"],
    "Incident Management": ["incident", "classification", "notification", "root cause", "response"],
    "Vendor / Third-Party Management": ["third-party", "third party", "vendor", "provider", "subcontract", "exit"],
    "Data Governance / Reporting": ["data", "metadata", "lineage", "dashboard", "report", "kpi", "kri"],
    "Internal Audit / Assurance": ["audit", "evidence", "traceability", "approval", "attestation"],
    "Operations / Settlement": ["operations", "settlement", "processing", "reconciliation", "back office"],
    "Programme Management": ["programme", "implementation", "roadmap", "budget", "sponsor", "maturity"],
    "Human Resources / Training": ["hr", "training", "people", "resource", "sme"],
}

# -----------------------------------------------------------------------------
# NOTE: All hardcoded option families, theme option dictionaries and
# implementation-status label tuples have been removed.
#
# Answer options and their scoring metadata are now generated per-question by
# the AI questionnaire agent (see services.ai_questionnaire_generator). Every
# option carries its own ``score_value``, ``readiness_interpretation``,
# ``triggers_followup`` flag and (optionally) an option-specific follow-up
# question that is directly tied to what the option reveals.
#
# Downstream code that used to read from ``DEFAULT_OPTIONS`` /
# ``THEME_OPTIONS`` / ``IMPLEMENTATION_STATUS_LABELS`` has been rewired to
# read from the per-option metadata attached to each question instead.
# -----------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Option helpers — options may now be plain strings OR dicts of the shape:
#   {"label": str, "next_question_ids": [str], "skip_question_ids": [str],
#    "score_value": int|float, "impact_value": ..., "risk_value": ...,
#    "branch_rule_id": str}
# Every helper below accepts both shapes so the rest of the pipeline is
# fully backward-compatible with the legacy ``List[str]`` form.
# ---------------------------------------------------------------------------

def option_label(option: Any) -> str:
    """Return the human-facing label for either a string option or a dict option."""
    if isinstance(option, Mapping):
        return str(option.get("label", "")).strip()
    return str(option).strip()


def option_labels(options: Sequence[Any]) -> List[str]:
    return [option_label(o) for o in options or []]


def option_metadata(options: Sequence[Any], label: str) -> Dict[str, Any]:
    """Return the option dict matching ``label``, or an empty dict for plain options."""
    for opt in options or []:
        if isinstance(opt, Mapping) and option_label(opt) == label:
            return dict(opt)
    return {}


# Legacy answer-to-score mapping. Kept as a *last-resort* fallback so that
# historical questionnaire packages (saved before the AI generator existed)
# can still be scored when they load from disk. New AI-generated packages
# carry the score on each option (``option.score_value``) — the scoring
# engine consults that first and only falls back to this table for legacy
# string-only option labels.
_LEGACY_ANSWER_SCORES = {
    "Yes": 100, "Complete": 100, "Measured / Optimised": 100, "Implemented": 90, "Mostly complete": 85,
    "Defined": 75, "Partially": 55, "Partially complete": 55, "Ad hoc": 35, "Not started": 0,
    "No": 0, "No owner assigned": 0, "No evidence available": 0, "Critical": 15, "High": 35,
    "Medium": 65, "Low": 90, "Named accountable owner": 95, "Shared ownership": 70,
    "Informal owner": 40, "Not applicable": None, "Unknown": 25,
    "Fully Implemented": 100,
    "Partially Implemented": 55,
    "Not Implemented": 0,
    "Not Applicable": None,
}

# Public alias retained so any legacy import (`from services.questionnaire_generator
# import ANSWER_SCORES`) keeps working; new code should not rely on this dict.
ANSWER_SCORES = _LEGACY_ANSWER_SCORES


# ---------------------------------------------------------------------------
# Domain dataclasses
# ---------------------------------------------------------------------------

@dataclass
class Requirement:
    source_section: str
    source_id: str
    normalized_id: str
    category: str
    requirement: str
    detail: str
    alignment: str
    priority: str
    acceptance: str
    confidence: int
    themes: List[str] = field(default_factory=list)
    #: Source citations attached to this requirement. Each entry is a plain
    #: dict with the ``SourceReference`` shape produced by
    #: :mod:`services.source_traceability`. Empty when no live source could
    #: be anchored to this requirement (UI surfaces this gap explicitly
    #: instead of fabricating a citation).
    source_references: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class ImpactPair:
    area: str
    function: str
    requirement_ids: List[str]
    regulatory_basis: str
    confidence: int


@dataclass
class Question:
    question_id: str
    area: str
    function: str
    question_type: str
    question: str
    options: List[Any]  # list of str OR list of dict (option metadata)
    mapped_requirement_ids: List[str]
    regulatory_basis: str
    confidence: int
    scoring_weight: int
    funnel_parent_id: str = ""
    trigger_answers: List[str] = field(default_factory=list)
    rationale: str = ""
    is_free_text: bool = False
    # --- v12 adaptive-branching extensions ------------------------------------
    # All new fields are optional and default-safe so older JSON packages still
    # load via ``Question(**q)`` (extra absent keys are fine; we never read them
    # without a default). The engine treats missing values as "use generic
    # routing" so behaviour is unchanged for legacy questions.
    branch_theme: str = ""          # e.g. "Incident reporting"
    branch_rule_id: str = ""        # set on dynamically-generated branch questions
    source_parent_id: str = ""      # the *original* base question that opened the branch
    dynamic_depth: int = 0          # 0 = base question, 1..N = dynamic depth
    # --- v13 explainability bundle -------------------------------------------
    # Structured traceability metadata so the UI and exports can answer
    # "Why am I being asked this?" without parsing free-form rationale text.
    # Schema (every key optional, missing keys treated as empty):
    #   regulation              str   (e.g. "DORA")
    #   regulator               str   (e.g. "ESMA / EBA")
    #   article                 str   (e.g. "DORA Article 17-20")
    #   obligation_id           str   (e.g. "OBL-014")
    #   brd_requirement_ids     [str] (e.g. ["BR-PRO-001", "BR-DAT-004"])
    #   rtm_trace_ids           [str] (e.g. ["TR-0014"])
    #   business_function       str   (e.g. "Incident Management")
    #   control_objective       str   (e.g. "Meet 24h initial notification window")
    #   reason                  str   (why this question exists)
    #   expected_evidence       str   (artefacts that would substantiate a positive answer)
    #   risk_if_negative        str   (consequence of a negative answer)
    #   owning_team             str   (Front Office / Compliance / Risk / ...)
    #   team_rationale          str   (why that team owns this question)
    #   impact_level            str   (Critical / High / Medium / Low)
    #   impact_reason           str   (why this impact level)
    #   plain_language_explainer str  (plain-English explanation of any reg jargon)
    #   evidence_expectations   [str] (concrete artefacts the client should produce)
    explainability: Dict[str, Any] = field(default_factory=dict)
    # --- AI-agent enhancement fields (per-question routing + impact) ---------
    # These are duplicated at the top level of the question dict for cheap
    # UI access (avoids reaching into ``explainability`` for every render).
    # Everything is optional / default-safe so legacy packages still load.
    owning_team: str = ""            # Front Office / Middle / Back / Risk / ...
    team_rationale: str = ""
    impact_level: str = ""           # Critical / High / Medium / Low
    impact_reason: str = ""
    impact_severity: str = ""        # kept in sync with impact_level for legacy readers
    impact_weight: int = 0           # numeric weight derived from impact severity
    plain_language_explainer: str = ""
    evidence_expectations: List[str] = field(default_factory=list)
    # Regulatory obligation IDs (Agent 1 OBL-XXX) the AI questionnaire
    # generator mapped this question to. Distinct from
    # ``mapped_requirement_ids`` which are BRD IDs.
    mapped_obligation_ids: List[str] = field(default_factory=list)
    # Parent / child relationships in the funnel tree. ``funnel_parent_id``
    # already captures the child->parent edge; ``child_question_ids`` is the
    # reverse edge, kept so the UI can walk the tree without a second pass.
    child_question_ids: List[str] = field(default_factory=list)
    is_parent: bool = False
    is_child: bool = False
    # True when the AI generator lacked enough context to produce a
    # grounded question and emitted an SME placeholder instead of
    # fabricating content.
    requires_manual_review: bool = False
    generated_by_ai: bool = False
    # Numeric priority derived from impact severity; used by the enhancer
    # to sort questions on the UI (higher = more urgent).
    priority_rank: int = 2
    # AI-agent question intent tagging: 'impact' / 'readiness' /
    # 'impact+readiness'. Populated by the AI questionnaire generator so
    # the UI + analytics can prove balanced impact-vs-readiness coverage
    # per impacted area.
    question_purpose: str = ""
    targets_impact_dimension: str = ""
    targets_readiness_dimension: str = ""
    # Populated by :mod:`services.question_style_enhancer` when the
    # question wording is intrinsically quantitative (Budget, Timeline,
    # Coverage %, Frequency, Team size, SLA). The scoring engine uses
    # this tag to award a small "quantitative depth" bonus in the
    # composite Evaluation Confidence score.
    quantitative_type: str = ""


# ---------------------------------------------------------------------------
# Helpers (lifted verbatim)
# ---------------------------------------------------------------------------

def clamp_confidence(value, floor: int = CONFIDENCE_FLOOR) -> int:
    try:
        score = int(str(value).replace("%", "").strip())
    except Exception:
        score = 94
    return max(floor, min(100, score))


def infer_themes(text: str) -> List[str]:
    lower = text.lower()
    themes: List[str] = []
    checks = [
        ("Governance", ["governance", "board", "management body", "approval"]),
        ("ICT risk management", ["ict risk", "framework", "risk management", "control"]),
        ("Incident reporting", ["incident", "classification", "notification", "reporting"]),
        ("Resilience testing", ["resilience", "testing", "backup", "restore", "recovery", "continuity"]),
        ("Third-party risk", ["third-party", "third party", "vendor", "provider", "contract", "subcontract", "exit"]),
        ("Data and evidence", ["data", "inventory", "lineage", "metadata", "evidence", "timestamp", "quality"]),
        ("Security and access", ["security", "access", "vulnerability", "encryption", "siem", "iam"]),
        ("Reporting", ["dashboard", "report", "kpi", "kri", "governance pack"]),
    ]
    for name, keys in checks:
        if any(k in lower for k in keys):
            themes.append(name)
    return themes or ["General regulatory coverage"]


def _section_prefix(section_label: str, source_id: str) -> str:
    section_l = (section_label or "").lower()
    source_l = (source_id or "").lower()
    if "process" in section_l or source_l.startswith("br-pro"):
        return "BR-PRO"
    if "data" in section_l or source_l.startswith("br-dat"):
        return "BR-DAT"
    if "report" in section_l or source_l.startswith("br-rep"):
        return "BR-REP"
    if "non-functional" in section_l or source_l.startswith("nfr"):
        return "NFR"
    if "functional" in section_l or source_l.startswith("fr"):
        return "FR"
    return "REQ"


# ---------------------------------------------------------------------------
# Requirement extraction from DOCX (now uses utils.docx_parser)
# ---------------------------------------------------------------------------

_REQUIRED_HEADERS = ("id", "category", "requirement", "detailedrequirement",
                     "doraalignment", "priority", "acceptancecriteria")


def read_docx_requirements(source: DocxSource) -> List[Requirement]:
    """Extract Requirement records from a BRD/FRD DOCX file or bytes.

    Equivalent to the original v11 ``read_docx_requirements`` but the
    document-iteration boilerplate has been replaced by a call to
    :func:`utils.docx_parser.iter_sectioned_tables`.
    """
    requirements: List[Requirement] = []
    counters: Dict[str, int] = defaultdict(int)
    for section_label, rows in iter_sectioned_tables(source):
        if not rows:
            continue
        headers = [normalise_header(h) for h in rows[0]]
        if not all(h in headers for h in _REQUIRED_HEADERS):
            continue
        idx = {h: i for i, h in enumerate(headers)}
        conf_idx = idx.get("aiconfidence")
        for row in rows[1:]:
            if len(row) < len(headers):
                row = list(row) + [""] * (len(headers) - len(row))
            source_id = row[idx["id"]]
            title = row[idx["requirement"]]
            detail = row[idx["detailedrequirement"]]
            if not (source_id or title or detail):
                continue
            prefix = _section_prefix(section_label, source_id)
            counters[prefix] += 1
            combined = " ".join([
                section_label, source_id, title, detail,
                row[idx["doraalignment"]], row[idx["acceptancecriteria"]],
            ])
            requirements.append(Requirement(
                source_section=section_label,
                source_id=source_id,
                normalized_id=f"{prefix}-{counters[prefix]:03d}",
                category=row[idx["category"]],
                requirement=title,
                detail=detail,
                alignment=row[idx["doraalignment"]],
                priority=row[idx["priority"]],
                acceptance=row[idx["acceptancecriteria"]],
                confidence=clamp_confidence(row[conf_idx] if conf_idx is not None else 95),
                themes=infer_themes(combined),
            ))
    return requirements


def requirements_from_report(
    report: Any,
    source_refs_by_item: Optional[Dict[str, List[Dict[str, Any]]]] = None,
) -> List[Requirement]:
    """Convert a Phase-4 DoraDetailedBRD into the Requirement list used here.

    Accepts the in-memory Pydantic ``DoraDetailedBRD`` model so the closed loop
    GenAI/offline -> questionnaire works without a DOCX round-trip.

    ``source_refs_by_item`` is the optional ``{REQ:<id> -> [SourceReference
    dict]}`` map produced by :func:`services.source_traceability.attach_source_references`
    and carried on ``BRDArtifact.metadata``. When supplied each returned
    ``Requirement`` carries the citations of the BRD row it was derived from
    so downstream UI / questionnaire / scoring layers can surface
    traceability without re-running the matcher.
    """
    section_map: List[Tuple[str, str, Any]] = [
        ("7.1 Process Requirements", "BR-PRO", report.process_business_requirements),
        ("7.2 Data Requirements", "BR-DAT", report.data_business_requirements),
        ("7.3 Reporting Requirements", "BR-REP", report.reporting_business_requirements),
        ("8. Functional Requirements", "FR", report.functional_requirements),
        ("10. Non-Functional Requirements", "NFR", report.non_functional_requirements),
    ]
    refs_map = source_refs_by_item or {}
    requirements: List[Requirement] = []
    counters: Dict[str, int] = defaultdict(int)
    for section_label, prefix, req_section in section_map:
        for item in req_section.items:
            counters[prefix] += 1
            combined = " ".join([
                section_label, item.id, item.category, item.requirement,
                item.detailed_requirement, item.regulation_alignment, item.acceptance_criteria,
            ])
            req_refs = list(refs_map.get(f"REQ:{item.id}", []))
            requirements.append(Requirement(
                source_section=section_label,
                source_id=item.id,
                normalized_id=f"{prefix}-{counters[prefix]:03d}",
                category=item.category,
                requirement=item.requirement,
                detail=item.detailed_requirement,
                alignment=item.regulation_alignment,
                priority=item.priority,
                acceptance=item.acceptance_criteria,
                confidence=clamp_confidence(item.confidence_level),
                themes=infer_themes(combined),
                source_references=req_refs,
            ))
    return requirements


# ---------------------------------------------------------------------------
# Impact-pair derivation
# ---------------------------------------------------------------------------

def score_keywords(text: str, keyword_map: Dict[str, List[str]]) -> Counter:
    lower = text.lower()
    scores: Counter = Counter()
    for label, keys in keyword_map.items():
        for key in keys:
            if key in lower:
                scores[label] += 1
    return scores


def impacted_labels_for_requirement(req: Requirement, keyword_map: Dict[str, List[str]], default: str) -> List[str]:
    text = " ".join([req.source_section, req.category, req.requirement, req.detail, req.alignment, req.acceptance])
    scores = score_keywords(text, keyword_map)
    if not scores:
        return [default]
    max_score = max(scores.values())
    selected = [k for k, v in scores.items() if v >= max(1, max_score - 1)]
    return selected[:3]


def regulatory_basis_for(reqs: Sequence[Requirement], regulation: str) -> str:
    alignments = [r.alignment for r in reqs if r.alignment]
    if alignments:
        return " | ".join(list(dict.fromkeys(alignments))[:3])
    return REGULATORY_TAXONOMY.get(regulation.upper(), {}).get("official_source", regulation)


def derive_impact_pairs(requirements: Sequence[Requirement], regulation: str) -> List[ImpactPair]:
    pair_to_ids: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    req_by_id = {r.normalized_id: r for r in requirements}
    for req in requirements:
        areas = impacted_labels_for_requirement(req, AREA_KEYWORDS, "Risk & Controls framework")
        functions = impacted_labels_for_requirement(req, FUNCTION_KEYWORDS, "Compliance & Legal")
        for area in areas:
            for function in functions:
                pair_to_ids[(area, function)].append(req.normalized_id)
    ranked = sorted(pair_to_ids.items(), key=lambda kv: (-len(set(kv[1])), kv[0][0], kv[0][1]))[:MAX_AREA_FUNCTION_PAIRS]
    pairs: List[ImpactPair] = []
    for (area, function), ids in ranked:
        unique_ids = list(dict.fromkeys(ids))
        reqs = [req_by_id[i] for i in unique_ids if i in req_by_id]
        avg_conf = round(sum(r.confidence for r in reqs) / max(1, len(reqs)))
        basis = regulatory_basis_for(reqs, regulation)
        pairs.append(ImpactPair(area, function, unique_ids, basis, clamp_confidence(avg_conf)))
    return pairs


# ---------------------------------------------------------------------------
# Question synthesis
# ---------------------------------------------------------------------------

# NOTE: ``select_option_family`` and ``_short_requirement_focus`` used to
# route hardcoded option families for a given requirement set. They have
# been removed. Option families are now generated per-question by the AI
# agent based on the specific question being asked (see
# services.ai_questionnaire_generator).


# ---------------------------------------------------------------------------
# Specificity helpers — extract concrete anchors from a requirement so that
# the generated questions reference the *actual* behaviour, metric, evidence
# expectation and regulatory article instead of generic placeholders.
# ---------------------------------------------------------------------------

# Regex kept for the content-correctness validator (see
# ``_cited_article_matches_requirement`` below), which cross-checks that an
# AI-generated question cites an article that actually appears in the
# mapped BRD row(s). This is a validation-only regex, not a template.
_ARTICLE_RE = _re.compile(
    r"\b(?:Article|Art\.?)\s*([0-9]{1,3}(?:\s*\([a-z0-9]+\))?(?:\s*\.\s*[0-9]+)?)\b",
    _re.IGNORECASE,
)

# Subject-verb stems that we strip from a behaviour clause so that what remains
# reads as a verb phrase (e.g. "monitor risks" rather than "The organization
# must monitor risks"). Order matters — longer phrases must come first.
# ---------------------------------------------------------------------------
# Hardcoded content removed: theme/option dictionaries, template question
# synthesisers, and free-text prompt families. Question generation is now
# 100% AI-driven (see services.ai_questionnaire_generator).
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------


def question_kind(question: Question) -> str:
    text = question.question.lower()
    # v13 — the canonical L1 status question. The plain-English rewrite
    # uses "Is the ... in place for ..." so we accept both the legacy
    # "implemented" wording and the new "in place" phrasing.
    if text.startswith("is the ") and ("implemented" in text or "in place" in text):
        return "coverage"
    if "implementation coverage" in text or "current implementation" in text:
        return "coverage"
    if (
        "named accountable owner" in text
        or "accountable owner" in text
        or "ownership model" in text
        or text.startswith("who owns ")
    ):
        return "ownership"
    if (
        "substantiating evidence" in text
        or "which evidence" in text
        or "evidence artefact" in text
        or "what evidence can" in text
        or text.startswith("what evidence")
    ):
        return "evidence"
    if (
        "remediation maturity" in text
        or ("remediation" in text and "gap" in text)
        or text.startswith("how mature is the remediation")
    ):
        return "remediation"
    if (
        "residual dora" in text
        or ("residual" in text and "risk" in text)
        or text.startswith("how much risk is still open")
    ):
        return "risk"
    # Theme deep-dive markers — each yields a distinct kind so the dedupe
    # logic does not collapse a third-party question with a security or
    # incident question for the same area/requirement set.
    if "incident" in text and ("classif" in text or "notification deadline" in text
                                or "report major incidents" in text):
        return "theme_incident"
    if (
        "third-party" in text
        or "third party" in text
        or "vendor contracts" in text
    ):
        return "theme_third_party"
    if (
        "resilience test" in text
        or "threat-led penetration" in text
        or "tlpt" in text
        or "resilience tests" in text
    ):
        return "theme_resilience"
    if (
        "privileged access" in text
        or "siem" in text
        or "encryption" in text
        or "security controls" in text
    ):
        return "theme_security"
    if ("management body" in text and "approv" in text) or "approved the policy" in text:
        return "theme_governance"
    if (
        "evidence dictionary" in text
        or "data lineage" in text
        or "metadata" in text
        or "evidence trail" in text
    ):
        return "theme_data"
    if (
        "management reporting" in text
        or "dashboard" in text
        or "kri" in text
        or "kpi" in text
    ):
        return "theme_reporting"
    if question.is_free_text:
        return "free_text"
    return _re.sub(r"[^a-z0-9]+", " ", text).strip()[:80]


def overlap_ratio(left: Sequence[str], right: Sequence[str]) -> float:
    a, b = set(left), set(right)
    if not a or not b:
        return 0.0
    return len(a & b) / min(len(a), len(b))


def dedupe_impact_pairs(pairs: Sequence[ImpactPair]) -> List[ImpactPair]:
    kept: List[ImpactPair] = []
    for pair in pairs:
        duplicate = False
        for existing in kept:
            same_axis = pair.area == existing.area or pair.function == existing.function
            if same_axis and overlap_ratio(pair.requirement_ids, existing.requirement_ids) >= 0.80:
                duplicate = True
                break
        if not duplicate:
            kept.append(pair)
    return kept


def dedupe_and_resequence_questions(questions: Sequence[Question]) -> List[Question]:
    kept: List[Question] = []
    seen_text: set[str] = set()
    seen_requirement_kind: set[Tuple[str, Tuple[str, ...]]] = set()
    seen_area_kind_primary: set[Tuple[str, str, str]] = set()

    for q in questions:
        norm_text = _re.sub(r"\W+", " ", q.question.lower()).strip()
        req_key = tuple(sorted(set(q.mapped_requirement_ids)))
        primary_req = req_key[0] if req_key else "FREE"
        kind = question_kind(q)

        requirement_kind_key = (kind, req_key)
        area_kind_primary_key = (kind, q.area, primary_req)

        if norm_text in seen_text:
            continue
        if not q.is_free_text and requirement_kind_key in seen_requirement_kind:
            continue
        if not q.is_free_text and area_kind_primary_key in seen_area_kind_primary:
            continue

        near_duplicate = False
        for existing in kept:
            if existing.is_free_text or q.is_free_text:
                continue
            same_kind = question_kind(existing) == kind
            same_axis = existing.area == q.area or existing.function == q.function
            req_overlap = overlap_ratio(existing.mapped_requirement_ids, q.mapped_requirement_ids)
            if same_kind and req_overlap >= 0.70 and same_axis:
                near_duplicate = True
                break
        if near_duplicate:
            continue

        seen_text.add(norm_text)
        seen_requirement_kind.add(requirement_kind_key)
        seen_area_kind_primary.add(area_kind_primary_key)
        kept.append(q)

    # Positional (non-severity) ordering: place every static child right
    # after its parent so the final Q-0001, Q-0002... numbering reflects the
    # natural funnel tree rather than an impact-severity re-sort. Roots keep
    # the order the AI produced them (impact pair by impact pair).
    kept_by_id: Dict[str, Question] = {q.question_id: q for q in kept}
    roots: List[Question] = []
    seen_roots: set = set()
    for q in kept:
        parent = (q.funnel_parent_id or "").strip()
        if not parent or parent not in kept_by_id:
            if q.question_id not in seen_roots:
                seen_roots.add(q.question_id)
                roots.append(q)

    ordered: List[Question] = []
    emitted_ids: set = set()

    def _emit(node: Question) -> None:
        if node.question_id in emitted_ids:
            return
        emitted_ids.add(node.question_id)
        ordered.append(node)
        for cid in node.child_question_ids or []:
            child = kept_by_id.get(cid)
            if child is not None:
                _emit(child)

    for root in roots:
        _emit(root)
    for q in kept:
        if q.question_id not in emitted_ids:
            emitted_ids.add(q.question_id)
            ordered.append(q)

    old_to_new: Dict[str, str] = {}
    for idx, q in enumerate(ordered, start=1):
        old_to_new[q.question_id] = f"Q-{idx:04d}"
    for idx, q in enumerate(ordered, start=1):
        q.question_id = f"Q-{idx:04d}"
    for q in ordered:
        if q.funnel_parent_id:
            new_parent = old_to_new.get(q.funnel_parent_id, "")
            q.funnel_parent_id = new_parent
            if not new_parent:
                q.trigger_answers = []
        if getattr(q, "source_parent_id", ""):
            q.source_parent_id = old_to_new.get(q.source_parent_id, q.source_parent_id)
        if q.child_question_ids:
            q.child_question_ids = [
                old_to_new[cid] for cid in q.child_question_ids if cid in old_to_new
            ]
        for opt in q.options or []:
            if isinstance(opt, dict) and opt.get("followup_question_id"):
                opt["followup_question_id"] = old_to_new.get(
                    opt["followup_question_id"], opt["followup_question_id"],
                )
    return ordered


def generate_question_bank(
    requirements: Sequence[Requirement],
    pairs: Sequence[ImpactPair],
    regulation: str = "DORA",
    *,
    obligations: Sequence[Any] = (),
    rtm_entries: Sequence[Any] = (),
    impact: Optional[Any] = None,
    readiness: Optional[Any] = None,
    client_roles: Sequence[str] = (),
    client_profile: Optional[Mapping[str, Any]] = None,
    source_refs_by_item: Optional[Mapping[str, List[Dict[str, Any]]]] = None,
    client: Optional[Any] = None,
) -> List[Question]:
    """Generate the questionnaire question bank via the AI agent.

    Delegates to :func:`services.ai_questionnaire_generator.generate_ai_questionnaire`.
    All hardcoded template functions have been removed from this module; the
    AI agent is the single source of question, option and follow-up content.

    The ``impact`` (Agent 1 :class:`ImpactAssessment`) and ``readiness``
    (Agent 1 :class:`ReadinessAssessment`) inputs are the two "context
    lenses" the AI generator uses to (a) target impact-probing questions
    at the specific affected items and (b) target readiness-probing
    questions at the client's weakest maturity dimensions.

    When ``client`` is ``None`` the AI generator emits a small set of "manual
    review required" placeholders (one per top impact pair) so downstream code
    still receives a valid package — it will NOT fabricate hardcoded questions.
    """
    from services.ai_questionnaire_generator import generate_ai_questionnaire

    clean_pairs = dedupe_impact_pairs(pairs)
    question_dicts = generate_ai_questionnaire(
        regulation=regulation,
        requirements=requirements,
        impact_pairs=clean_pairs,
        obligations=obligations,
        rtm_entries=rtm_entries,
        impact=impact,
        readiness=readiness,
        client_roles=client_roles,
        client_profile=client_profile,
        source_refs_by_item=source_refs_by_item,
        client=client,
    )
    questions: List[Question] = []
    for q_dict in question_dicts:
        try:
            filtered = {
                k: v for k, v in q_dict.items()
                if k in Question.__dataclass_fields__
            }
            questions.append(Question(**filtered))
        except Exception:
            continue
    questions = dedupe_and_resequence_questions(questions)
    # Style diversification: turn suitable questions into Multi Select and
    # inject quantitative bracket options for Budget / Timeline / Coverage /
    # Frequency / Team size / SLA wording. Deterministic + idempotent.
    from services.question_style_enhancer import (
        diversify_question_styles,
        sanitize_questions_without_options,
    )
    diversify_question_styles(questions)
    # Never surface a closed question without answerable options - the SME
    # would just see an empty "— Select an answer —" dropdown. Free-text
    # style wording is auto-converted to Open Ended, anything else is
    # dropped. Runs after diversification so we don't accidentally strip
    # newly-injected quantitative bracket options.
    questions = sanitize_questions_without_options(questions)
    questions = dedupe_and_resequence_questions(questions)
    # Guarantee that every closed L1 question has at least one adaptive
    # follow-up child. The AI system prompt already requires this, but
    # LLMs occasionally forget - the deterministic guarantee injects a
    # synthetic evidence follow-up tied to the option with the lowest
    # readiness score so the questionnaire always feels adaptive.
    questions = ensure_funnel_followups(questions)
    return questions


def ensure_funnel_followups(questions: List[Question]) -> List[Question]:
    """Guarantee every closed L1 question exposes at least one funnel child.

    Rules:
      * Free-text and child questions are left untouched.
      * If a closed L1 question already has ``child_question_ids`` populated
        (via the LLM) it is left untouched.
      * Otherwise, we pick the option with the lowest ``score_value`` (or
        the first option when scores are missing), mark it as
        ``triggers_followup=True`` with a fresh ``followup_question_id``
        pointing at a new free-text evidence question. The new question
        is appended to the list, then the whole list is re-numbered so
        IDs stay contiguous.

    Idempotent: running twice does not add extra follow-ups because the
    first pass populates ``child_question_ids`` which short-circuits the
    check on the second pass.
    """
    if not questions:
        return questions

    parents_needing_children: List[Question] = []
    for q in questions:
        if q.is_free_text or q.is_child or q.funnel_parent_id:
            continue
        if q.child_question_ids:
            continue
        # Only closed questions with usable options can spawn a follow-up.
        opts = q.options or []
        if not opts:
            continue
        parents_needing_children.append(q)

    if not parents_needing_children:
        return questions

    next_index = len(questions) + 1

    def _option_label(opt: Any) -> str:
        if isinstance(opt, Mapping):
            return str(opt.get("label") or opt.get("value") or "").strip()
        return str(opt or "").strip()

    def _option_score(opt: Any) -> Optional[float]:
        if isinstance(opt, Mapping):
            for key in ("score_value", "readiness_score", "score", "value"):
                v = opt.get(key)
                if isinstance(v, (int, float)):
                    return float(v)
        return None

    for parent in parents_needing_children:
        opts = list(parent.options or [])
        # Find the option with the lowest readiness score (weakest state)
        # so the follow-up probes the biggest gap. If no options carry a
        # numeric score, fall back to the first option.
        weak_idx = 0
        weak_score = float("inf")
        for i, opt in enumerate(opts):
            sv = _option_score(opt)
            if sv is None:
                continue
            if sv < weak_score:
                weak_score = sv
                weak_idx = i
        weak_opt = opts[weak_idx]
        weak_label = _option_label(weak_opt) or "the answer above"

        child_qid = f"Q-{next_index:04d}"
        next_index += 1

        # Mutate the parent's option (dict shape only) so the UI knows to
        # branch on this option. Plain-string options get promoted to a
        # dict in-place.
        if isinstance(weak_opt, Mapping):
            weak_opt_dict = dict(weak_opt)
        else:
            weak_opt_dict = {"label": str(weak_opt), "score_value": None}
        weak_opt_dict["triggers_followup"] = True
        weak_opt_dict["followup_question_id"] = child_qid
        opts[weak_idx] = weak_opt_dict
        parent.options = opts
        parent.child_question_ids = list(parent.child_question_ids or []) + [child_qid]
        parent.is_parent = True

        # Compose the child evidence question. Deterministic wording so
        # every follow-up reads consistently.
        child_question_text = (
            f"You indicated \"{weak_label}\" for the previous question. "
            f"Describe the current state, any evidence in place, and the "
            f"specific gap that would need to close for this to move to a "
            f"stronger rating."
        )
        child_rationale = (
            f"Adaptive follow-up triggered because the previous answer "
            f"(\"{weak_label}\") indicates a state that requires more "
            f"detail to be defensible in an audit."
        )

        child_explainability = dict(parent.explainability or {})
        child_explainability["reason"] = child_rationale
        child_explainability["question_purpose"] = (
            parent.explainability.get("question_purpose")
            if isinstance(parent.explainability, Mapping) else "readiness"
        ) or "readiness"

        child = Question(
            question_id=child_qid,
            area=parent.area,
            function=parent.function,
            question_type="Open Ended",
            question=child_question_text,
            options=[],
            mapped_requirement_ids=list(parent.mapped_requirement_ids or []),
            regulatory_basis=parent.regulatory_basis,
            confidence=max(70, parent.confidence - 10),
            scoring_weight=max(1, parent.scoring_weight),
            funnel_parent_id=parent.question_id,
            trigger_answers=[weak_label],
            rationale=child_rationale,
            is_free_text=True,
            branch_theme=parent.branch_theme or parent.area,
            source_parent_id=parent.question_id,
            dynamic_depth=1,
            explainability=child_explainability,
            owning_team=parent.owning_team,
            team_rationale=parent.team_rationale,
            impact_level=parent.impact_level,
            impact_reason=parent.impact_reason,
            impact_severity=parent.impact_severity,
            impact_weight=parent.impact_weight,
            plain_language_explainer=parent.plain_language_explainer,
            evidence_expectations=list(parent.evidence_expectations or []),
            mapped_obligation_ids=list(parent.mapped_obligation_ids or []),
            is_child=True,
            requires_manual_review=False,
            generated_by_ai=False,
            priority_rank=parent.priority_rank,
        )
        questions.append(child)

    return questions


# ---------------------------------------------------------------------------
# Validation, scoring, package, evaluation
# ---------------------------------------------------------------------------

_STOPWORDS = frozenset((
    "the", "and", "for", "with", "from", "into", "this", "that", "must",
    "shall", "should", "will", "have", "has", "are", "was", "were", "been",
    "their", "there", "such", "any", "all", "only", "also", "more", "than",
    "may", "can", "via", "per", "incl", "including", "use", "used", "using",
    "based", "ensure", "ensures", "ensuring", "applicable", "relevant",
    "where", "which", "what", "when", "how", "etc", "ict",
))


def _token_bag(text: Optional[str]) -> set:
    """Lowercased lemma-ish word bag, stopword-filtered, length > 3."""
    if not text:
        return set()
    raw = _re.sub(r"[^a-zA-Z0-9\s]", " ", str(text)).lower().split()
    return {t for t in raw if len(t) > 3 and t not in _STOPWORDS}


def _cited_article_matches_requirement(question_article: str, mapped_reqs: Sequence[Requirement]) -> bool:
    """True iff the article cited by the question appears in at least one mapped requirement's alignment.

    Catches hallucinated citations — a question that says "DORA Article 30"
    when its mapped BRD row only talks about Article 6 fails this check.
    """
    if not question_article or not mapped_reqs:
        return False
    q_match = _ARTICLE_RE.search(question_article)
    if q_match:
        q_num = q_match.group(1).strip().lower()
        for req in mapped_reqs:
            for source in (req.alignment, req.detail, req.requirement, req.acceptance):
                if not source:
                    continue
                for r_match in _ARTICLE_RE.findall(source):
                    if r_match.strip().lower() == q_num:
                        return True
        return False
    # No article number in the question citation — fall back to literal substring match.
    needle = question_article.lower().strip()
    return any(needle in (req.alignment or "").lower() for req in mapped_reqs)


def _question_text_anchors_to_requirement(question_text: str, mapped_reqs: Sequence[Requirement]) -> bool:
    """True iff the question text shares at least 3 content tokens with the mapped requirement."""
    if not question_text or not mapped_reqs:
        return False
    q_tokens = _token_bag(question_text)
    if not q_tokens:
        return False
    for req in mapped_reqs:
        req_tokens = _token_bag(" ".join([
            req.requirement or "", req.detail or "", req.acceptance or "",
            req.category or "",
        ]))
        if len(q_tokens & req_tokens) >= 3:
            return True
    return False


def _evidence_anchors_to_acceptance(expected_evidence: str, mapped_reqs: Sequence[Requirement]) -> bool:
    """True iff the question's expected-evidence shares vocabulary with the mapped acceptance criteria."""
    if not expected_evidence or not mapped_reqs:
        return False
    ev_tokens = _token_bag(expected_evidence)
    if not ev_tokens:
        return False
    for req in mapped_reqs:
        accept_tokens = _token_bag(req.acceptance) | _token_bag(req.detail)
        if len(ev_tokens & accept_tokens) >= 2:
            return True
    return False


def _question_cites_specific_article(question_text: str, rationale: str) -> bool:
    """True iff the question or its rationale cites a specific article number (not just 'DORA')."""
    blob = " ".join([question_text or "", rationale or ""])
    return bool(_ARTICLE_RE.search(blob))


def _explainability_is_complete(explainability: Mapping[str, Any]) -> bool:
    """True iff every required explainability key is populated and non-empty."""
    if not explainability:
        return False
    for key in EXPLAINABILITY_REQUIRED_KEYS:
        value = explainability.get(key)
        if value is None or value == "" or value == [] or value == {}:
            return False
    return True


def _l1_option_is_grounded(q: Question) -> bool:
    """True iff an L1 root question's options carry per-option branch metadata.

    Returns False for non-root questions (they have a funnel parent) so this
    only counts L1 status questions in the denominator.
    """
    if q.funnel_parent_id or q.is_free_text:
        return False
    opts = q.options or []
    if not opts or not isinstance(opts[0], Mapping):
        return False
    dict_opts = [o for o in opts if isinstance(o, Mapping)]
    if not dict_opts:
        return False
    return all(
        o.get("branch_rule_id") and "score_value" in o
        for o in dict_opts
    )


def _validate_structural_completeness(
    requirements: Sequence[Requirement],
    pairs: Sequence[ImpactPair],
    questions: Sequence[Question],
) -> Tuple[float, Dict[str, float]]:
    """Legacy v11 structural-completeness grade.

    Kept verbatim from v11/v12 so audit trails comparing v13.1 packages to
    historical ones remain like-for-like. Returns ``(overall_pct, breakdown)``.
    """
    req_ids = {r.normalized_id for r in requirements}
    mapped = {rid for q in questions for rid in q.mapped_requirement_ids}
    req_cov = len(req_ids & mapped) / max(1, len(req_ids))
    areas = {p.area for p in pairs}
    q_areas = {q.area for q in questions if not q.is_free_text}
    area_cov = len(areas & q_areas) / max(1, len(areas))
    functions = {p.function for p in pairs}
    q_functions = {q.function for q in questions if not q.is_free_text}
    fn_cov = len(functions & q_functions) / max(1, len(functions))
    avg_q_conf = sum(q.confidence for q in questions) / max(1, len(questions))
    free_text_count = sum(1 for q in questions if q.is_free_text)
    free_text_ok = 1 if MIN_FREE_TEXT <= free_text_count <= MAX_FREE_TEXT else 0
    pair_depth = min(1, (len([q for q in questions if not q.is_free_text]) / max(1, len(pairs) * 4)))
    overall = round(
        (0.30 * req_cov + 0.18 * area_cov + 0.18 * fn_cov + 0.14 * pair_depth + 0.10 * free_text_ok + 0.10 * (avg_q_conf / 100)) * 100,
        1,
    )
    breakdown = {
        "requirement_coverage_pct": round(req_cov * 100, 1),
        "area_coverage_pct": round(area_cov * 100, 1),
        "function_coverage_pct": round(fn_cov * 100, 1),
        "pair_depth_pct": round(pair_depth * 100, 1),
        "average_question_confidence_pct": round(avg_q_conf, 1),
        "free_text_question_count": free_text_count,
        "free_text_in_target_band": bool(free_text_ok),
    }
    return overall, breakdown


def _validate_content_correctness(
    requirements: Sequence[Requirement],
    questions: Sequence[Question],
) -> Tuple[float, Dict[str, float]]:
    """v13.1 content-correctness grade.

    Tests whether each generated question is **grounded** in its mapped BRD
    row(s):

    * cites the right article (no hallucinated citations),
    * shares vocabulary with the requirement detail / acceptance,
    * references the BRD's expected-evidence phrasing,
    * carries a complete explainability bundle,
    * traces back through ``mapped_requirement_ids`` + ``obligation_id``,
    * cites a specific article number (not just the regulation name),
    * (for L1 roots only) has dict-shaped options with per-option
      ``branch_rule_id`` + ``score_value`` so the adaptive funnel can route.

    Returns ``(overall_pct, breakdown)``. The overall is a weighted sum of
    seven per-question pass-rates; weights sum to 1.0.
    """
    closed = [q for q in questions if not q.is_free_text]
    if not closed:
        return 0.0, {
            "article_citation_match_pct": 0.0,
            "behaviour_anchoring_pct": 0.0,
            "evidence_anchoring_pct": 0.0,
            "traceability_completeness_pct": 0.0,
            "explainability_completeness_pct": 0.0,
            "specificity_pct": 0.0,
            "l1_option_grounding_pct": 0.0,
            "closed_question_count": 0,
            "l1_root_question_count": 0,
        }

    req_by_id = {r.normalized_id: r for r in requirements}
    article_hits = 0
    anchor_hits = 0
    evidence_hits = 0
    traceability_hits = 0
    explainability_hits = 0
    specificity_hits = 0
    l1_grounded_hits = 0
    l1_total = 0

    for q in closed:
        mapped_reqs = [req_by_id[rid] for rid in q.mapped_requirement_ids if rid in req_by_id]
        explainability = q.explainability or {}
        question_article = (
            str(explainability.get("article") or q.regulatory_basis or "")
        )
        if _cited_article_matches_requirement(question_article, mapped_reqs):
            article_hits += 1
        if _question_text_anchors_to_requirement(q.question, mapped_reqs):
            anchor_hits += 1
        expected_evidence = str(explainability.get("expected_evidence") or "")
        if _evidence_anchors_to_acceptance(expected_evidence, mapped_reqs):
            evidence_hits += 1
        if q.mapped_requirement_ids and explainability.get("obligation_id"):
            traceability_hits += 1
        if _explainability_is_complete(explainability):
            explainability_hits += 1
        if _question_cites_specific_article(q.question, q.rationale):
            specificity_hits += 1
        if not q.funnel_parent_id:
            l1_total += 1
            if _l1_option_is_grounded(q):
                l1_grounded_hits += 1

    n = len(closed)
    article_rate = article_hits / n
    anchor_rate = anchor_hits / n
    evidence_rate = evidence_hits / n
    traceability_rate = traceability_hits / n
    explainability_rate = explainability_hits / n
    specificity_rate = specificity_hits / n
    l1_rate = (l1_grounded_hits / l1_total) if l1_total else 1.0

    # Weights sum to 1.0
    overall = (
        0.25 * article_rate
        + 0.20 * explainability_rate
        + 0.15 * traceability_rate
        + 0.15 * anchor_rate
        + 0.10 * specificity_rate
        + 0.10 * evidence_rate
        + 0.05 * l1_rate
    ) * 100

    breakdown = {
        "article_citation_match_pct": round(article_rate * 100, 1),
        "behaviour_anchoring_pct": round(anchor_rate * 100, 1),
        "evidence_anchoring_pct": round(evidence_rate * 100, 1),
        "traceability_completeness_pct": round(traceability_rate * 100, 1),
        "explainability_completeness_pct": round(explainability_rate * 100, 1),
        "specificity_pct": round(specificity_rate * 100, 1),
        "l1_option_grounding_pct": round(l1_rate * 100, 1),
        "closed_question_count": n,
        "l1_root_question_count": l1_total,
    }
    return round(overall, 1), breakdown


def validate_and_score_package(
    requirements: Sequence[Requirement],
    pairs: Sequence[ImpactPair],
    questions: Sequence[Question],
) -> Tuple[int, Dict[str, Any]]:
    """Compute the package-confidence headline + a full audit breakdown.

    The headline number returned (``overall``) is:

    * **Content-correctness grade** (default, ``PACKAGE_CONFIDENCE_MODE=content``)
      — seven groundedness signals weighted to test whether each question is
      anchored in its BRD row. See :func:`_validate_content_correctness`.
    * **Structural-completeness grade** (legacy, ``PACKAGE_CONFIDENCE_MODE=structural``)
      — the v11/v12 formula testing whether all BRD requirements, areas,
      functions and pairs are covered. See :func:`_validate_structural_completeness`.

    Both grades are always computed and surfaced in ``metrics`` so audits and
    the recommendations engine can read either.
    """
    content_pct, content_breakdown = _validate_content_correctness(requirements, questions)
    structural_pct, structural_breakdown = _validate_structural_completeness(
        requirements, pairs, questions,
    )

    if PACKAGE_CONFIDENCE_MODE == "structural":
        headline = structural_pct
        mode_used = "structural"
    else:
        headline = content_pct
        mode_used = "content"

    headline = max(OVERALL_CONFIDENCE_FLOOR, min(100, headline))
    metrics: Dict[str, Any] = {
        "package_confidence_mode": mode_used,
        "content_correctness_pct": round(content_pct, 1),
        "structural_completeness_pct": round(structural_pct, 1),
        "content_breakdown": content_breakdown,
        "structural_breakdown": structural_breakdown,
        # Back-compat keys read by older dashboards / exports.
        "requirement_coverage_pct": structural_breakdown["requirement_coverage_pct"],
        # ``coverage_pct`` is the headline "Coverage (closed)" KPI shown on
        # Page 3: what fraction of BRD requirements have at least one closed
        # (mapped) question. Same value as ``requirement_coverage_pct``;
        # kept as an explicit alias because the cockpit reads this key.
        "coverage_pct": structural_breakdown["requirement_coverage_pct"],
        "area_coverage_pct": structural_breakdown["area_coverage_pct"],
        "function_coverage_pct": structural_breakdown["function_coverage_pct"],
        "pair_depth_pct": structural_breakdown["pair_depth_pct"],
        "average_question_confidence_pct": structural_breakdown["average_question_confidence_pct"],
        "free_text_question_count": structural_breakdown["free_text_question_count"],
        "closed_question_count": content_breakdown["closed_question_count"],
        "question_count": len(questions),
        "requirement_count": len(requirements),
        "impact_pair_count": len(pairs),
    }
    return int(round(headline)), metrics


def evaluate_responses(questions: Sequence[Question], responses: Mapping[str, Any]) -> Dict[str, Any]:
    """Deterministic batch evaluation (kept here for backwards compatibility).

    The live Streamlit cockpit uses an adaptive variant exposed through
    :mod:`services.scoring_engine` in Phase 6. The two implementations share
    :data:`ANSWER_SCORES`.
    """
    numerator = 0.0
    denominator = 0.0
    by_area: Dict[str, List[float]] = defaultdict(lambda: [0.0, 0.0])
    details: List[Dict[str, Any]] = []
    for q in questions:
        if q.is_free_text:
            continue
        raw = responses.get(q.question_id)
        values = raw if isinstance(raw, list) else [raw]
        scores: List[float] = []
        for value in values:
            if value in ANSWER_SCORES and ANSWER_SCORES[value] is not None:
                scores.append(float(ANSWER_SCORES[value]))  # type: ignore[arg-type]
        if not scores:
            score = 25.0 if raw not in (None, "", []) else 0.0
        else:
            score = sum(scores) / len(scores)
        weighted = score * q.scoring_weight * (q.confidence / 100)
        max_weighted = 100 * q.scoring_weight * (q.confidence / 100)
        numerator += weighted
        denominator += max_weighted
        by_area[q.area][0] += weighted
        by_area[q.area][1] += max_weighted
        details.append({
            "question_id": q.question_id,
            "score": round(score, 1),
            "weighted_score": round(weighted, 1),
            "max_weighted": round(max_weighted, 1),
        })
    compliance = round((numerator / denominator) * 100, 1) if denominator else 0.0
    confidence = round(min(99, max(90, sum(q.confidence for q in questions) / max(1, len(questions)) - (5 if not responses else 0))), 1)
    return {
        "compliance_score_pct": compliance,
        "evaluation_confidence_pct": confidence,
        "area_scores": {area: round(vals[0] / vals[1] * 100, 1) if vals[1] else 0 for area, vals in by_area.items()},
        "details": details,
        "explanation": (
            "Score is weighted by regulatory priority/scoring weight and discounted by per-question confidence. "
            "Not applicable responses are excluded where possible; unknown/unanswered responses reduce confidence and score."
        ),
    }


def package_dict(
    regulation: str,
    requirements: Sequence[Requirement],
    pairs: Sequence[ImpactPair],
    questions: Sequence[Question],
    overall_confidence: int,
    metrics: Dict[str, float],
) -> Dict[str, Any]:
    mode = (metrics or {}).get("package_confidence_mode", PACKAGE_CONFIDENCE_MODE)
    note = (
        "Content-correctness grade: every question is tested for article-citation match, "
        "behaviour/evidence anchoring against the BRD, traceability completeness, "
        "explainability bundle completeness, specificity, and (for L1 roots) "
        "per-option branch grounding. Deterministic; not legal advice."
        if mode == "content"
        else
        "Structural-completeness grade (legacy): tests whether all BRD requirements, "
        "impacted areas/functions and pairs have questions. Deterministic; not legal advice."
    )
    # Promote a handful of headline metrics from ``metrics`` to the top of
    # ``metadata`` so the cockpit can read them directly without diving into
    # the nested ``metrics`` blob. ``overall_confidence_pct`` was always
    # surfaced this way; ``coverage_pct`` is the headline "Coverage (closed)"
    # tile on Page 3 (fraction of BRD requirements covered by at least one
    # closed/mapped question).
    coverage_pct = (
        metrics.get("coverage_pct")
        if metrics
        else None
    )
    if coverage_pct is None and metrics:
        coverage_pct = metrics.get("requirement_coverage_pct", 0)
    return {
        "metadata": {
            "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "regulation": regulation,
            "overall_confidence_pct": overall_confidence,
            "coverage_pct": coverage_pct if coverage_pct is not None else 0,
            "confidence_note": note,
            "metrics": metrics,
            "regulatory_taxonomy": REGULATORY_TAXONOMY.get(regulation.upper(), {}),
        },
        "requirements": [asdict(r) for r in requirements],
        "impact_pairs": [asdict(p) for p in pairs],
        "questions": [asdict(q) for q in questions],
        "answer_scores": ANSWER_SCORES,
    }


# ---------------------------------------------------------------------------
# Top-level orchestrators
# ---------------------------------------------------------------------------

def _build_package(
    requirements: List[Requirement],
    regulation: str,
    *,
    obligations: Sequence[Any] = (),
    rtm_entries: Sequence[Any] = (),
    impact: Optional[Any] = None,
    readiness: Optional[Any] = None,
    client_roles: Sequence[str] = (),
    client_profile: Optional[Mapping[str, Any]] = None,
    source_refs_by_item: Optional[Mapping[str, List[Dict[str, Any]]]] = None,
    client: Optional[Any] = None,
) -> Dict[str, Any]:
    """Assemble a full questionnaire package (AI-driven).

    ``client`` is the optional :class:`~services.genai_service.GenAIClient`.
    When present the AI generator produces the full question bank; when
    absent the AI generator emits manual-review placeholders (no
    hardcoded template questions are used).

    ``impact`` and ``readiness`` are the Agent 1 assessments; when
    provided the AI generator uses them to (a) target impact-probing
    questions at the specific affected items and (b) target
    readiness-probing questions at the weakest maturity dimensions.
    """
    logger.info(
        "Building questionnaire package. regulation=%s requirements=%d obligations=%d roles=%s client=%s",
        regulation, len(requirements), len(obligations or []),
        list(client_roles or []) or None,
        "genai" if client is not None else "offline",
    )
    pairs = dedupe_impact_pairs(derive_impact_pairs(requirements, regulation))
    logger.info("Impact pairs derived. total=%d (after dedup)", len(pairs))
    questions = generate_question_bank(
        requirements, pairs, regulation=regulation,
        obligations=obligations,
        rtm_entries=rtm_entries,
        impact=impact,
        readiness=readiness,
        client_roles=client_roles,
        client_profile=client_profile,
        source_refs_by_item=source_refs_by_item,
        client=client,
    )
    overall, metrics = validate_and_score_package(requirements, pairs, questions)
    logger.info(
        "Questionnaire package assembled. questions=%d overall_confidence=%d",
        len(questions), overall,
    )
    return package_dict(regulation, requirements, pairs, questions, overall, metrics)


def build_questionnaire_package(
    source: DocxSource,
    regulation: str = "DORA",
    *,
    obligations: Sequence[Any] = (),
    rtm_entries: Sequence[Any] = (),
    impact: Optional[Any] = None,
    readiness: Optional[Any] = None,
    client_roles: Sequence[str] = (),
    client_profile: Optional[Mapping[str, Any]] = None,
    client: Optional[Any] = None,
) -> Dict[str, Any]:
    """Parse a BRD/FRD DOCX and return the complete questionnaire package.

    Optional context (obligations / rtm_entries / impact / readiness /
    client_roles / client_profile / client) is forwarded to the AI
    questionnaire agent so the generated questions are scoped to the
    selected client type and grounded in the live regulatory obligations
    and Agent 1 impact + readiness assessments.
    """
    requirements = read_docx_requirements(source)
    if not requirements:
        raise ValueError(
            "No BRD/FRD requirement tables were found. Ensure the DOCX contains tables with "
            "the columns: ID, Category, Requirement, Detailed Requirement, DORA Alignment, "
            "Priority, Acceptance Criteria (AI Confidence optional)."
        )
    return _build_package(
        requirements, regulation,
        obligations=obligations,
        rtm_entries=rtm_entries,
        impact=impact,
        readiness=readiness,
        client_roles=client_roles,
        client_profile=client_profile,
        client=client,
    )


def build_package_from_report(
    report: Any,
    regulation: str = "DORA",
    source_refs_by_item: Optional[Dict[str, List[Dict[str, Any]]]] = None,
    *,
    obligations: Sequence[Any] = (),
    rtm_entries: Sequence[Any] = (),
    impact: Optional[Any] = None,
    readiness: Optional[Any] = None,
    client_roles: Sequence[str] = (),
    client_profile: Optional[Mapping[str, Any]] = None,
    client: Optional[Any] = None,
) -> Dict[str, Any]:
    """Build a questionnaire package directly from a Phase-4 DoraDetailedBRD model.

    This is the closed-loop path used when the Streamlit Page 1 "Generate BRD/FRD
    from regulation" option produces an in-memory report.

    ``source_refs_by_item`` is forwarded to :func:`requirements_from_report`
    so every question inherits the citations of the BRD requirement it was
    derived from. Pass the ``source_references_by_item`` map carried on
    ``BRDArtifact.metadata`` to enable end-to-end traceability into the
    questionnaire.

    The AI-agent context parameters (``obligations``, ``rtm_entries``,
    ``impact``, ``readiness``, ``client_roles``, ``client_profile``,
    ``client``) are forwarded to :mod:`services.ai_questionnaire_generator`
    so the questions are grounded in the live regulatory analysis, RTM
    entries, Agent 1 impact + readiness assessments and selected client
    type.
    """
    requirements = requirements_from_report(report, source_refs_by_item or {})
    if not requirements:
        raise ValueError("Report has no requirements; cannot build questionnaire package.")
    return _build_package(
        requirements, regulation,
        obligations=obligations,
        rtm_entries=rtm_entries,
        impact=impact,
        readiness=readiness,
        client_roles=client_roles,
        client_profile=client_profile,
        source_refs_by_item=source_refs_by_item,
        client=client,
    )


# ---------------------------------------------------------------------------
# Excel writer (lifted verbatim from v11)
# ---------------------------------------------------------------------------

def write_excel(
    path: str,
    requirements: Sequence[Requirement],
    pairs: Sequence[ImpactPair],
    questions: Sequence[Question],
    overall_confidence: int,
    metrics: Dict[str, float],
) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    impact = wb.create_sheet("Impacted Functions Areas")
    qsheet = wb.create_sheet("Questionnaire")
    free = wb.create_sheet("Free Text Questions")
    funnel = wb.create_sheet("Funnel Logic")
    scoring = wb.create_sheet("Scoring Rubric")
    trace = wb.create_sheet("Requirement Traceability")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(sheet):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = white
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    ws.append(["Metric", "Value"])
    for key, value in metrics.items():
        if isinstance(value, Mapping):
            ws.append([f"-- {key} --", ""])
            for sub_key, sub_value in value.items():
                ws.append([f"   {sub_key}", sub_value])
        else:
            ws.append([key, value])
    ws.append(["Overall Confidence", f"{overall_confidence}%"])
    ws.append(["Important note",
               "Confidence is a model/rule assurance indicator and not a legal opinion. "
               "Compliance and Legal should validate before client use."])
    style_sheet(ws)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 80

    impact.append(["Impact Area", "Function", "Mapped Requirement IDs", "Regulatory Basis", "Confidence"])
    for p in pairs:
        impact.append([p.area, p.function, " | ".join(p.requirement_ids), p.regulatory_basis, p.confidence])
    style_sheet(impact)
    for col, width in enumerate([36, 32, 42, 58, 14], 1):
        impact.column_dimensions[get_column_letter(col)].width = width

    headers = [
        "Question ID", "Impact Area", "Function", "Type", "Question", "Options",
        "Mapped Requirement IDs", "Regulatory Basis", "Confidence", "Scoring Weight",
        "Funnel Parent", "Trigger Answers", "Rationale",
    ]
    qsheet.append(headers)
    free.append(headers)
    for q in questions:
        rendered_options = " | ".join(option_label(o) for o in q.options)
        row = [q.question_id, q.area, q.function, q.question_type, q.question, rendered_options,
               " | ".join(q.mapped_requirement_ids), q.regulatory_basis, q.confidence, q.scoring_weight,
               q.funnel_parent_id, " | ".join(q.trigger_answers), q.rationale]
        (free if q.is_free_text else qsheet).append(row)
    style_sheet(qsheet)
    style_sheet(free)
    widths = [14, 32, 28, 14, 70, 48, 38, 46, 14, 14, 14, 36, 70]
    for sheet in [qsheet, free]:
        for col, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        for r in range(2, sheet.max_row + 1):
            sheet.row_dimensions[r].height = 58

    funnel.append(["Question ID", "Parent Question ID", "Trigger Answers", "Funnel Rule"])
    for q in questions:
        rule = "Always visible" if not q.funnel_parent_id else f"Show when {q.funnel_parent_id} in {', '.join(q.trigger_answers)}"
        funnel.append([q.question_id, q.funnel_parent_id, " | ".join(q.trigger_answers), rule])
    style_sheet(funnel)
    for col, width in enumerate([16, 20, 44, 80], 1):
        funnel.column_dimensions[get_column_letter(col)].width = width

    scoring.append(["Answer", "Score", "Interpretation"])
    for answer, score in ANSWER_SCORES.items():
        scoring.append([answer, "Excluded" if score is None else score,
                        "Mapped deterministic score used by Streamlit evaluation"])
    style_sheet(scoring)
    scoring.column_dimensions["A"].width = 34
    scoring.column_dimensions["B"].width = 14
    scoring.column_dimensions["C"].width = 60

    trace.append(["Requirement ID", "Source ID", "Section", "Category", "Requirement",
                  "DORA Alignment", "Priority", "Mapped Question IDs", "Mapped Question Count"])
    for r in requirements:
        qids = [q.question_id for q in questions if r.normalized_id in q.mapped_requirement_ids]
        trace.append([r.normalized_id, r.source_id, r.source_section, r.category, r.requirement,
                      r.alignment, r.priority, " | ".join(qids), len(qids)])
    style_sheet(trace)
    for col, width in enumerate([18, 18, 32, 24, 54, 42, 12, 62, 18], 1):
        trace.column_dimensions[get_column_letter(col)].width = width

    if qsheet.max_row > 1:
        dv = DataValidation(
            type="list",
            formula1='"Yes,Partially,No,Complete,Mostly complete,Partially complete,Not started,Unknown,Not applicable"',
            allow_blank=True,
        )
        qsheet.add_data_validation(dv)
    wb.save(path)
    return os.path.abspath(path)


def _filter_dataclass_kwargs(cls: Any, payload: Mapping[str, Any]) -> Dict[str, Any]:
    """Drop unknown keys so packages serialised by a future/older version still load."""
    allowed = {f.name for f in cls.__dataclass_fields__.values()}  # type: ignore[attr-defined]
    return {k: v for k, v in payload.items() if k in allowed}


def write_excel_from_package(path: str, package: Mapping[str, Any]) -> str:
    """Convenience writer that hydrates the package dict back into dataclasses first."""
    requirements = [Requirement(**_filter_dataclass_kwargs(Requirement, r)) for r in package["requirements"]]
    pairs = [ImpactPair(**_filter_dataclass_kwargs(ImpactPair, p)) for p in package["impact_pairs"]]
    questions = [Question(**_filter_dataclass_kwargs(Question, q)) for q in package["questions"]]
    meta = package.get("metadata", {})
    overall = int(meta.get("overall_confidence_pct", OVERALL_CONFIDENCE_FLOOR))
    metrics = dict(meta.get("metrics", {}))
    return write_excel(path, requirements, pairs, questions, overall, metrics)


__all__ = [
    # Legacy scoring compatibility (used only when loading pre-AI packages).
    "ANSWER_SCORES",
    # Impact taxonomies used by regulatory_analysis_agent to bucket
    # requirements — NOT question templates.
    "AREA_KEYWORDS",
    "FUNCTION_KEYWORDS",
    "CONFIDENCE_FLOOR",
    "EXPLAINABILITY_REQUIRED_KEYS",
    "ImpactPair",
    "option_label",
    "option_labels",
    "option_metadata",
    "MAX_AREA_FUNCTION_PAIRS",
    "MAX_FREE_TEXT",
    "MIN_FREE_TEXT",
    "OVERALL_CONFIDENCE_FLOOR",
    "PACKAGE_CONFIDENCE_MODE",
    "Question",
    "REGULATORY_TAXONOMY",
    "Requirement",
    "build_package_from_report",
    "build_questionnaire_package",
    "clamp_confidence",
    "dedupe_and_resequence_questions",
    "dedupe_impact_pairs",
    "derive_impact_pairs",
    "ensure_funnel_followups",
    "evaluate_responses",
    "generate_question_bank",
    "impacted_labels_for_requirement",
    "infer_themes",
    "package_dict",
    "question_kind",
    "read_docx_requirements",
    "regulatory_basis_for",
    "requirements_from_report",
    "score_keywords",
    "validate_and_score_package",
    "write_excel",
    "write_excel_from_package",
]
